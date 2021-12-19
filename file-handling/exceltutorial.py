#first we create an empty excel file then we stored it in file variable, then we accessed the active sheet 
from openpyxl import workbook   #woekbook is a function which returns an excel file
file= Workbook()	#nameless workbook created- calling workbook fn
wst=file.active	    # active sheet selected and stored it in wst
wst['A1']=88		# writing in the cell keys= cell value
wst.title='mysheet'
file.create_sheet('school')	# creating a new sheet by the name school
teachers=file['school']	#/  # teachers=file['school']	# 2 ways to access the sheet by the name school	
teachers['B5']= 'Riya'
file.save('marks.xlsx')		#giving the excel its name

Accessing multiple cells at one time
teachers.cell(row=9,column=4, value=99)		#another way of adding data to a cell
file.save('marks.xlsx')	#saves the file with changes that we hav made abv, file name is what we have passed as argument

#Accessing multiple cells at one time
# for i in range (1,11):
# 	teachers.cell(row=i,column=5,value=i)
# num=11
# for i in range (1,11):
# 	teachers.cell(row=i,column=6,value=num)
# 	num=num+1
# file.save('marks.xlsx')

# fil=open('file.txt','w')	#if we want to create a file then we have to open it in write/apend mode
# file_new=fil.write()
# print(file_new)
# fil.close()

file=open('myfile.txt','w')
file.write('happy')
file.close()

file=open('myfile.txt')
result=file.read()
print(result)

# myfile=open('docs.txt','a')
#  =myfile.

accessing a pre existing excel file-
from openpyxl import load_workbook
file=load_workbook('marks.xlsx')
print(file.sheetnames)
